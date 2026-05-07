"""
로또 당첨번호 이력 분석 스크립트
- 입력: lotto_history.xlsx (C2:I열에 최신회차부터 당첨번호6개+보너스번호)
- 출력: 동일 파일에 M, N~T, U열 추가
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = "lotto_history.xlsx"
OUTPUT_FILE = "lotto_history_analyzed.xlsx"

wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb.active

# ── 데이터 읽기 ──────────────────────────────────────────────
# C=3, D=4, E=5, F=6, G=7, H=8, I=9 (당첨번호6 + 보너스1)
# 행2부터 데이터 시작, 최신회차가 위에 있음

rows_data = []  # [(n1,n2,n3,n4,n5,n6,bonus), ...]  인덱스0=최신
row = 2
while True:
    vals = [ws.cell(row=row, column=c).value for c in range(3, 10)]
    if vals[0] is None:
        break
    rows_data.append(tuple(int(v) for v in vals))
    row += 1

total = len(rows_data)  # 전체 회차 수

# ── 스타일 정의 ──────────────────────────────────────────────
hdr_fill  = PatternFill("solid", start_color="1F4E79")
hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
ng_fill   = PatternFill("solid", start_color="FF9999")
dup_fill  = PatternFill("solid", start_color="FFD966")
num_font  = Font(name="Arial", size=9)
center    = Alignment(horizontal="center", vertical="center")
thin      = Side(style="thin", color="CCCCCC")
border    = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_cell(cell, value, fill=None, bold=False, number_format=None):
    cell.value = value
    cell.font = Font(bold=bold, name="Arial", size=9,
                     color="FFFFFF" if fill == hdr_fill else "000000")
    cell.alignment = center
    cell.border = border
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format

# ── 헤더 작성 ────────────────────────────────────────────────
# M열(13): 이전5회차 중복합
# N~T열(14~20): 7개 번호별 누적 당첨횟수
# U열(21): 분산 (또는 NG)

style_cell(ws.cell(row=1, column=13), "이전5회차\n중복합", hdr_fill)
ws.row_dimensions[1].height = 28

num_labels = ["번호1", "번호2", "번호3", "번호4", "번호5", "번호6", "보너스"]
for i, lbl in enumerate(num_labels):
    style_cell(ws.cell(row=1, column=14+i), f"{lbl}\n누적횟수", hdr_fill)

style_cell(ws.cell(row=1, column=21), "분산\n(NG=중복)", hdr_fill)

# 열 너비 조정
ws.column_dimensions["M"].width = 10
ws.column_dimensions["U"].width = 10
for col in range(14, 21):
    ws.column_dimensions[get_column_letter(col)].width = 9

# ── M열: 이전 5회차 내 중복 번호 합산 ────────────────────────
# rows_data[0]=최신(엑셀 row2), rows_data[i] → excel row = i+2
# 이전5회차 = rows_data[i+1] ~ rows_data[i+5]
# 6회차(index5) 이상부터 계산 가능

dup_counts = []  # M열 값 저장 (인덱스 = 회차 인덱스)

for i in range(total):
    excel_row = i + 2
    if i + 5 >= total:
        # 이전 5회차가 존재하지 않으면 공란
        ws.cell(row=excel_row, column=13).value = None
        dup_counts.append(None)
        continue

    current_set = set(rows_data[i])  # 현재 회차 7개 번호
    prev_numbers = set()
    for j in range(i+1, i+6):
        prev_numbers.update(rows_data[j])

    dup = len(current_set & prev_numbers)
    dup_counts.append(dup)

    cell_m = ws.cell(row=excel_row, column=13)
    style_cell(cell_m, dup, dup_fill if dup > 0 else None, number_format="0")

# ── N~U열: 누적 당첨 횟수 및 분산 ───────────────────────────
# 각 회차 i에서:
#   해당 회차(i)까지만 사용 → rows_data[i:]  (최신이 위이므로 i~total-1이 해당 회차 이전)
#   즉, 엑셀 행 기준으로 row2가 1000회차면 rows_data[0..999]가 1회~1000회
#   rows_data[i]는 (total-i)번째 회차 → 해당 회차 포함 이전 = rows_data[i:]

# 전체 번호 풀(1~45) 각 번호 등장 횟수를 슬라이딩으로 계산
# 아래쪽(rows_data[i:])이 해당 회차 이전 전체 이력

# 누적 카운트: 아래부터 누적 (bottom-up)
from collections import defaultdict

# running_count[i] = rows_data[i]까지 포함한 누적 카운트 (i부터 끝까지)
# 역순으로 순회
running_count = defaultdict(int)
cumulative = [None] * total  # cumulative[i] = dict {번호: 횟수} (rows_data[i..total-1])

for i in range(total - 1, -1, -1):
    for num in rows_data[i]:
        running_count[num] += 1
    cumulative[i] = dict(running_count)

# 전체 번호(1~45)의 총 평균 당첨횟수 계산 함수
def calc_variance(cnt_dict, total_draws):
    """
    total_draws개 회차에서 1~45 번호별 등장횟수의 분산
    등장횟수 평균 = total_draws * 7 / 45
    분산 = mean((count_i - mean)^2) for i in 1..45
    """
    expected_draws = total_draws  # 회차 수
    total_appearances = expected_draws * 7  # 총 번호 등장 수
    avg = total_appearances / 45
    variance = sum((cnt_dict.get(n, 0) - avg) ** 2 for n in range(1, 46)) / 45
    return round(variance, 4)

for i in range(total):
    excel_row = i + 2
    cnt_dict = cumulative[i]  # rows_data[i]부터 끝까지 누적
    draws_count = total - i   # 해당 회차 포함 이전 전체 회차 수

    # N~T: 현재 회차 7개 번호 각각의 누적 당첨 횟수
    for j, num in enumerate(rows_data[i]):
        cell = ws.cell(row=excel_row, column=14+j)
        style_cell(cell, cnt_dict.get(num, 0), number_format="0")

    # U열: 분산 또는 NG
    cell_u = ws.cell(row=excel_row, column=21)
    dup = dup_counts[i]

    if dup is None:
        # 이전 5회차 미존재 → 분산만 표시
        var = calc_variance(cnt_dict, draws_count)
        style_cell(cell_u, var, number_format="0.0000")
    elif dup > 0:
        # 중복 있음 → NG
        style_cell(cell_u, "NG", ng_fill)
    else:
        var = calc_variance(cnt_dict, draws_count)
        style_cell(cell_u, var, number_format="0.0000")

# 나머지 기존 셀 스타일 정비 (M열 None인 경우 빈칸 처리)
for i in range(total):
    if dup_counts[i] is None:
        excel_row = i + 2
        ws.cell(row=excel_row, column=13).value = "-"
        ws.cell(row=excel_row, column=13).alignment = center
        ws.cell(row=excel_row, column=13).border = border

wb.save(OUTPUT_FILE)
print(f"완료: {OUTPUT_FILE} 저장됨 ({total}개 회차 처리)")
